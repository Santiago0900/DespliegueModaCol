from django.http import HttpResponse
from django.conf import settings
from datetime import datetime
import os

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from infrastructure.persistence.repositories.django_cliente_repository import DjangoClienteRepository
from infrastructure.persistence.repositories.django_categoria_repository import DjangoCategoriaRepository
from infrastructure.persistence.repositories.django_usuario_repository import DjangoUsuarioRepository
from infrastructure.persistence.repositories.django_producto_repository import DjangoProductoRepository
from infrastructure.persistence.repositories.django_proveedor_repository import DjangoProveedorRepository
from infrastructure.persistence.repositories.django_flujo_caja_repository import DjangoFlujoCajaRepository
from infrastructure.persistence.repositories.django_rol_repository import DjangoRolRepository
from infrastructure.persistence.repositories.django_venta_repository import DjangoVentaRepository
from infrastructure.persistence.repositories.django_compra_repository import DjangoCompraRepository


def reporte_general(request):
    tipo = request.GET.get("tipo", "clientes")
    formato = request.GET.get("formato", "pdf")
    activo = request.GET.get("activo")
    buscar = request.GET.get("buscar")

    if tipo == "compras":
        repo = DjangoCompraRepository()
        items = repo.listar()

        fecha_compra = request.GET.get("fechaCompra")
        fecha_entrega = request.GET.get("fechaEstimadaEntrega")
        proveedor_id = request.GET.get("proveedorId")

        if fecha_compra:
            items = [c for c in items if str(c.fecha_compra) == fecha_compra]

        if fecha_entrega:
            items = [c for c in items if str(c.fecha_estimada_entrega) == fecha_entrega]

        if proveedor_id:
            items = [c for c in items if str(c.proveedor_id) == str(proveedor_id)]

        if formato == "excel":
            return generar_excel_compras(items)
        return generar_pdf_compras(items)


    if tipo == "roles":
        repo = DjangoRolRepository()
        items = repo.listar()

        if formato == "excel":
            return generar_excel_roles(items)
        return generar_pdf_roles(items)

    if tipo == "ventas":
        repo = DjangoVentaRepository()
        items = repo.listar()

        if formato == "excel":
            return generar_excel_ventas(items)
        return generar_pdf_ventas(items)

    # -------- FLUJO DE CAJA --------
    if tipo == "flujo-caja":
        repo = DjangoFlujoCajaRepository()

        # aceptar varios nombres de parámetros por compatibilidad
        desde = (
                request.GET.get("desde")
                or request.GET.get("fechaInicio")
                or request.GET.get("fecha")
                or None
        )
        hasta = (
                request.GET.get("hasta")
                or request.GET.get("fechaFin")
                or None
        )
        tipo_mov = (
                request.GET.get("tipo_mov")
                or request.GET.get("tipoMovimiento")
                or request.GET.get("tipo")
                or None
        )

        # evitar conflicto: si viene tipo=flujo-caja en la URL, no usarlo como tipo_movimiento
        if tipo_mov == "flujo-caja":
            tipo_mov = None

        items = repo.filtrar(desde, hasta, tipo_mov)

        if formato == "excel":
            return generar_excel_flujo_caja(items, desde, hasta)
        return generar_pdf_flujo_caja(items, desde, hasta)


    # -------- PROVEEDORES --------
    if tipo == "proveedores":
        repo = DjangoProveedorRepository()
        items = repo.listar()

        if activo is not None:
            activo_bool = activo.lower() == "true"
            items = [p for p in items if p.activo == activo_bool]

        if buscar:
            buscar = buscar.lower()
            items = [
                p for p in items
                if buscar in p.razon_social.lower()
                   or buscar in p.identificacion.lower()
                   or (p.correo and buscar in p.correo.lower())
            ]

        if formato == "excel":
            return generar_excel_proveedores(items)
        return generar_pdf_proveedores(items)

    # -------- PRODUCTOS --------
    if tipo == "productos":
        repo = DjangoProductoRepository()
        items = repo.listar()

        if activo is not None:
            activo_bool = activo.lower() == "true"
            items = [p for p in items if p.activo == activo_bool]

        if buscar:
            buscar = buscar.lower()
            items = [
                p for p in items
                if buscar in p.nombre.lower()
                   or (p.descripcion and buscar in p.descripcion.lower())
                   or (p.proveedor_nombre and buscar in p.proveedor_nombre.lower())
            ]

        if formato == "excel":
            return generar_excel_productos(items)
        return generar_pdf_productos(items)

    # -------- CATEGORÍAS --------
    if tipo == "categorias":
        repo = DjangoCategoriaRepository()
        items = repo.listar()

        if activo is not None:
            activo_bool = activo.lower() == "true"
            items = [c for c in items if c.activo == activo_bool]

        if buscar:
            buscar = buscar.lower()
            items = [c for c in items if buscar in c.tipo_categoria.lower()]

        if formato == "excel":
            return generar_excel_categorias(items)
        return generar_pdf_categorias(items)

    # -------- USUARIOS --------
    if tipo == "usuarios":
        repo = DjangoUsuarioRepository()
        items = repo.listar()

        if activo is not None:
            activo_bool = activo.lower() == "true"
            items = [u for u in items if u.activo == activo_bool]

        if buscar:
            buscar = buscar.lower()
            items = [
                u for u in items
                if buscar in u.nombre.lower()
                   or buscar in u.correo.lower()
                   or (u.rol_tipo and buscar in u.rol_tipo.lower())
            ]

        if formato == "excel":
            return generar_excel_usuarios(items)
        return generar_pdf_usuarios(items)

    # -------- DASHBOARD --------
    if tipo == "dashboard":
        usuario_repo = DjangoUsuarioRepository()
        cliente_repo = DjangoClienteRepository()
        producto_repo = DjangoProductoRepository()
        proveedor_repo = DjangoProveedorRepository()
        venta_repo = DjangoVentaRepository()
        flujo_repo = DjangoFlujoCajaRepository()

        hoy = datetime.now().date()
        inicio_mes = hoy.replace(day=1)

        fecha_inicio = request.GET.get("fechaInicio") or inicio_mes.strftime("%Y-%m-%d")
        fecha_fin = request.GET.get("fechaFin") or hoy.strftime("%Y-%m-%d")

        ventas = venta_repo.filtrar(fecha_inicio, fecha_fin, None)
        movimientos = flujo_repo.filtrar(fecha_inicio, fecha_fin)

        dashboard_data = {
            "total_usuarios": len([u for u in usuario_repo.listar() if getattr(u, "activo", True)]),
            "total_clientes": len([c for c in cliente_repo.listar() if getattr(c, "activo", True)]),
            "total_productos": len([p for p in producto_repo.listar() if getattr(p, "activo", True)]),
            "total_proveedores": len([p for p in proveedor_repo.listar() if getattr(p, "activo", True)]),
            "ventas_periodo": len(ventas),
            "ingresos_periodo": sum((v.total or 0) for v in ventas),
            "egresos_periodo": sum((m.monto or 0) for m in movimientos if m.tipo_movimiento == "EGRESO"),
            "saldo_periodo": (
                    sum((m.monto or 0) for m in movimientos if m.tipo_movimiento == "INGRESO")
                    - sum((m.monto or 0) for m in movimientos if m.tipo_movimiento == "EGRESO")
            ),
            "productos_stock_bajo": len([
                p for p in producto_repo.listar()
                if getattr(p, "activo", True) and (p.cantidad or 0) < 10
            ]),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "ultimas_ventas": sorted(
                venta_repo.listar(),
                key=lambda v: (v.fecha_venta or datetime.min.date(), v.id or 0),
                reverse=True
            )[:10],
        }

        if formato == "excel":
            return generar_excel_dashboard(dashboard_data)
        return generar_pdf_dashboard(dashboard_data)

    # -------- CLIENTES (default) --------
    repo = DjangoClienteRepository()
    clientes = repo.listar()

    if activo is not None:
        activo_bool = activo.lower() == "true"
        clientes = [c for c in clientes if c.activo == activo_bool]

    if buscar:
        buscar = buscar.lower()
        clientes = [
            c for c in clientes
            if buscar in c.nombre.lower()
               or buscar in c.empresa.lower()
               or buscar in c.identificacion.lower()
        ]

    if formato == "excel":
        return generar_excel_clientes(clientes)

    return generar_pdf_clientes(clientes)

def generar_pdf_compras(compras):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_compras.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Compras", styles["Title"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["Número", "Fecha", "Proveedor", "Usuario", "Estado", "Total"]]
    for c in compras:
        data.append([
            c.numero_compra,
            str(c.fecha_compra) if c.fecha_compra else "",
            c.proveedor_nombre or "",
            c.usuario_nombre or "",
            c.estado,
            str(c.total),
            ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_compras(compras):
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    _estilizar_excel(ws, "REPORTE DE COMPRAS", ["Número", "Fecha", "Proveedor", "Usuario", "Estado", "Total"])

    for c in compras:
        ws.append([
            c.numero_compra,
            str(c.fecha_compra) if c.fecha_compra else "",
            c.proveedor_nombre or "",
            c.usuario_nombre or "",
            c.estado,
            float(c.total),
            ])

    _ajustar_excel(ws, "A4:F4")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_compras.xlsx"
    wb.save(response)
    return response


def generar_pdf_roles(roles):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_roles.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Roles", styles["Title"]))
    elements.append(Spacer(1, 10))

    data = [["ID", "Tipo", "Estado"]]
    for r in roles:
        data.append([str(r.id), r.tipo, "Activo" if r.activo else "Inactivo"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#000000")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_roles(roles):
    wb = Workbook()
    ws = wb.active
    ws.title = "Roles"
    _estilizar_excel(ws, "REPORTE DE ROLES", ["ID", "Tipo", "Estado"])

    for r in roles:
        ws.append([r.id, r.tipo, "Activo" if r.activo else "Inactivo"])

    _ajustar_excel(ws, "A4:C4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_roles.xlsx"
    wb.save(response)
    return response


def generar_pdf_ventas(ventas):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Ventas", styles["Title"]))
    elements.append(Spacer(1, 10))

    data = [["ID", "Fecha", "Cliente", "Usuario", "Total", "Estado"]]
    for v in ventas:
        data.append([
            str(v.id),
            str(v.fecha_venta) if v.fecha_venta else "",
            v.cliente_nombre or "",
            v.usuario_nombre or "",
            str(v.total),
            v.estado,
            ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#000000")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_ventas(ventas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    _estilizar_excel(ws, "REPORTE DE VENTAS", ["ID", "Fecha", "Cliente", "Usuario", "Total", "Estado"])

    for v in ventas:
        ws.append([
            v.id,
            str(v.fecha_venta) if v.fecha_venta else "",
            v.cliente_nombre or "",
            v.usuario_nombre or "",
            float(v.total),
            v.estado,
            ])

    _ajustar_excel(ws, "A4:F4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_ventas.xlsx"
    wb.save(response)
    return response

def agregar_logo(elements):
    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logo.png")
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=120, height=50))
        elements.append(Spacer(1, 10))


def generar_pdf_flujo_caja(movimientos, desde=None, hasta=None):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_flujo_caja.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Flujo de Caja", styles["Title"]))
    rango = f"Rango: {desde or 'Inicio'} - {hasta or 'Hoy'}"
    elements.append(Paragraph(rango, styles["Normal"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["Fecha", "Tipo", "Descripción", "Monto", "Origen"]]
    for m in movimientos:
        origen = "Manual"
        if m.venta_id:
            origen = f"Venta #{m.venta_id}"
        elif m.compra_id:
            origen = f"Compra #{m.compra_id}"

        data.append([
            str(m.fecha) if m.fecha else "",
            m.tipo_movimiento,
            m.descripcion or "",
            str(m.monto),
            origen,
            ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_flujo_caja(movimientos, desde=None, hasta=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "FlujoCaja"
    _estilizar_excel(ws, "REPORTE DE FLUJO DE CAJA", ["Fecha", "Tipo", "Descripción", "Monto", "Origen"])

    for m in movimientos:
        origen = "Manual"
        if m.venta_id:
            origen = f"Venta #{m.venta_id}"
        elif m.compra_id:
            origen = f"Compra #{m.compra_id}"

        ws.append([
            str(m.fecha) if m.fecha else "",
            m.tipo_movimiento,
            m.descripcion or "",
            float(m.monto),
            origen,
            ])

    _ajustar_excel(ws, "A4:E4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_flujo_caja.xlsx"
    wb.save(response)
    return response


def generar_pdf_proveedores(proveedores):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_proveedores.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Proveedores", styles["Title"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["ID", "Razón Social", "Identificación", "Correo", "Contacto", "Estado"]]
    for p in proveedores:
        data.append([
            str(p.id),
            p.razon_social,
            p.identificacion,
            p.correo or "",
            p.contacto or "",
            "Activo" if p.activo else "Inactivo",
            ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_proveedores(proveedores):
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    _estilizar_excel(ws, "REPORTE DE PROVEEDORES", ["ID", "Razón Social", "Identificación", "Correo", "Contacto", "Estado"])

    for p in proveedores:
        ws.append([
            p.id,
            p.razon_social,
            p.identificacion,
            p.correo or "",
            p.contacto or "",
            "Activo" if p.activo else "Inactivo",
            ])

    _ajustar_excel(ws, "A4:F4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_proveedores.xlsx"
    wb.save(response)
    return response


def generar_pdf_productos(productos):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_productos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Productos", styles["Title"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["ID", "Fecha", "Nombre", "Categoría", "Precio", "Cantidad", "Proveedor", "Estado"]]
    for p in productos:
        data.append([
            str(p.id),
            str(p.fecha) if p.fecha else "",
            p.nombre,
            p.categoria_nombre or "",
            str(p.precio_unitario),
            str(p.cantidad),
            p.proveedor_nombre or "",
            "Activo" if p.activo else "Inactivo",
            ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_productos(productos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    _estilizar_excel(
        ws,
        "REPORTE DE PRODUCTOS",
        ["ID", "Fecha", "Nombre", "Categoría", "Precio", "Cantidad", "Proveedor", "Estado"]
    )

    for p in productos:
        ws.append([
            p.id,
            str(p.fecha) if p.fecha else "",
            p.nombre,
            p.categoria_nombre or "",
            float(p.precio_unitario),
            p.cantidad,
            p.proveedor_nombre or "",
            "Activo" if p.activo else "Inactivo",
            ])

    _ajustar_excel(ws, "A4:H4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_productos.xlsx"
    wb.save(response)
    return response


def generar_pdf_clientes(clientes):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_clientes.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Clientes", styles["Title"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["ID", "Empresa", "Nombre", "Identificación", "Contacto", "Correo", "Estado"]]
    for c in clientes:
        data.append([
            str(c.id),
            c.empresa,
            c.nombre,
            c.identificacion,
            c.contacto or "",
            c.correo or "",
            "Activo" if c.activo else "Inactivo",
            ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_clientes(clientes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    _estilizar_excel(ws, "REPORTE DE CLIENTES", ["ID", "Empresa", "Nombre", "Identificación", "Contacto", "Correo", "Estado"])

    for c in clientes:
        ws.append([
            c.id,
            c.empresa,
            c.nombre,
            c.identificacion,
            c.contacto or "",
            c.correo or "",
            "Activo" if c.activo else "Inactivo"
        ])

    _ajustar_excel(ws, "A4:G4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_clientes.xlsx"
    wb.save(response)
    return response


def generar_pdf_categorias(categorias):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_categorias.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Categorías", styles["Title"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["ID", "Tipo de Categoría", "Estado"]]
    for c in categorias:
        data.append([
            str(c.id),
            c.tipo_categoria,
            "Activo" if c.activo else "Inactivo",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_categorias(categorias):
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    _estilizar_excel(ws, "REPORTE DE CATEGORÍAS", ["ID", "Tipo de Categoría", "Estado"])

    for c in categorias:
        ws.append([c.id, c.tipo_categoria, "Activo" if c.activo else "Inactivo"])

    _ajustar_excel(ws, "A4:C4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_categorias.xlsx"
    wb.save(response)
    return response


def generar_pdf_usuarios(usuarios):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_usuarios.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte de Usuarios", styles["Title"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["ID", "Nombre", "Correo", "Rol", "Estado"]]
    for u in usuarios:
        data.append([
            str(u.id),
            u.nombre,
            u.correo,
            u.rol_tipo or "",
            "Activo" if u.activo else "Inactivo",
            ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


def generar_excel_usuarios(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    _estilizar_excel(ws, "REPORTE DE USUARIOS", ["ID", "Nombre", "Correo", "Rol", "Estado"])

    for u in usuarios:
        ws.append([
            u.id,
            u.nombre,
            u.correo,
            u.rol_tipo or "",
            "Activo" if u.activo else "Inactivo",
            ])

    _ajustar_excel(ws, "A4:E4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_usuarios.xlsx"
    wb.save(response)
    return response


def generar_pdf_dashboard(data):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_dashboard.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    agregar_logo(elements)
    elements.append(Paragraph("Reporte Completo del Dashboard", styles["Title"]))
    elements.append(Paragraph(
        f"Periodo: {data['fecha_inicio']} a {data['fecha_fin']}",
        styles["Normal"]
    ))
    elements.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 10))

    resumen = [
        ["Métrica", "Valor"],
        ["Total Usuarios", str(data["total_usuarios"])],
        ["Total Clientes", str(data["total_clientes"])],
        ["Total Productos", str(data["total_productos"])],
        ["Total Proveedores", str(data["total_proveedores"])],
        ["Ventas del Periodo", str(data["ventas_periodo"])],
        ["Ingresos del Periodo", str(data["ingresos_periodo"])],
        ["Egresos del Periodo", str(data["egresos_periodo"])],
        ["Saldo del Periodo", str(data["saldo_periodo"])],
        ["Productos con Stock Bajo", str(data["productos_stock_bajo"])],
    ]

    table_resumen = Table(resumen, repeatRows=1)
    table_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elements.append(table_resumen)
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("Últimas Ventas", styles["Heading2"]))
    elements.append(Spacer(1, 6))

    ventas_data = [["ID", "Fecha", "Cliente", "Usuario", "Total", "Estado"]]
    for v in data["ultimas_ventas"]:
        ventas_data.append([
            str(v.id),
            str(v.fecha_venta) if v.fecha_venta else "",
            v.cliente_nombre or "",
            v.usuario_nombre or "",
            str(v.total),
            v.estado or "",
            ])

    table_ventas = Table(ventas_data, repeatRows=1)
    table_ventas.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1f1f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
    ]))

    elements.append(table_ventas)
    doc.build(elements)
    return response


def generar_excel_dashboard(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    _estilizar_excel(ws, "REPORTE COMPLETO DEL DASHBOARD", ["Métrica", "Valor"])

    resumen = [
        ["Periodo", f"{data['fecha_inicio']} a {data['fecha_fin']}"],
        ["Total Usuarios", data["total_usuarios"]],
        ["Total Clientes", data["total_clientes"]],
        ["Total Productos", data["total_productos"]],
        ["Total Proveedores", data["total_proveedores"]],
        ["Ventas del Periodo", data["ventas_periodo"]],
        ["Ingresos del Periodo", float(data["ingresos_periodo"])],
        ["Egresos del Periodo", float(data["egresos_periodo"])],
        ["Saldo del Periodo", float(data["saldo_periodo"])],
        ["Productos con Stock Bajo", data["productos_stock_bajo"]],
    ]

    for item in resumen:
        ws.append(item)

    ws.append([])
    ws.append(["ÚLTIMAS VENTAS"])
    ws.append(["ID", "Fecha", "Cliente", "Usuario", "Total", "Estado"])

    for v in data["ultimas_ventas"]:
        ws.append([
            v.id,
            str(v.fecha_venta) if v.fecha_venta else "",
            v.cliente_nombre or "",
            v.usuario_nombre or "",
            float(v.total),
            v.estado or "",
            ])

    _ajustar_excel(ws, "A4:B4")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=reporte_dashboard.xlsx"
    wb.save(response)
    return response


def _estilizar_excel(ws, titulo, headers):
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    end_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = titulo
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center_align

    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = center_align

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border


def _ajustar_excel(ws, filtro_ref):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3

    ws.auto_filter.ref = filtro_ref
    ws.freeze_panes = "A5"